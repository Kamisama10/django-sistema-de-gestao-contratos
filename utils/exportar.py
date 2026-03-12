# utils/exportar.py

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import io

# ── Paleta ────────────────────────────────────────────────────────────────────
PRETO      = colors.HexColor('#111111')
CINZA_ESC  = colors.HexColor('#444444')
CINZA_MED  = colors.HexColor('#888888')
CINZA_BORD = colors.HexColor('#dddddd')
CINZA_BG   = colors.HexColor('#f7f7f7')
BRANCO     = colors.white
AZUL       = colors.HexColor('#1d4ed8')
VERDE      = colors.HexColor('#16a34a')
VERDE_BG   = colors.HexColor('#f0fdf4')
VERMELHO   = colors.HexColor('#dc2626')
VERM_BG    = colors.HexColor('#fff5f5')
AMAR_BG    = colors.HexColor('#fffbeb')


def _header_pdf(canvas, doc, titulo, subtitulo=''):
    canvas.saveState()
    w, h = doc.pagesize

    # Linha azul topo — 3px
    canvas.setFillColor(AZUL)
    canvas.rect(0, h - 3, w, 3, fill=1, stroke=0)

    # Logo
    canvas.setFillColor(PRETO)
    canvas.setFont('Helvetica-Bold', 11)
    canvas.drawString(1.8*cm, h - 22, 'Mobilizze')
    canvas.setFillColor(AZUL)
    canvas.drawString(1.8*cm + 52, h - 22, '.')

    # Título direita
    canvas.setFillColor(CINZA_ESC)
    canvas.setFont('Helvetica', 9)
    canvas.drawRightString(w - 1.8*cm, h - 19, titulo)
    if subtitulo:
        canvas.setFillColor(CINZA_MED)
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(w - 1.8*cm, h - 30, subtitulo)

    # Linha separadora
    canvas.setStrokeColor(CINZA_BORD)
    canvas.setLineWidth(0.5)
    canvas.line(1.8*cm, h - 38, w - 1.8*cm, h - 38)

    # Rodapé
    canvas.line(1.8*cm, 1.4*cm, w - 1.8*cm, 1.4*cm)
    canvas.setFillColor(CINZA_MED)
    canvas.setFont('Helvetica', 7.5)
    canvas.drawString(1.8*cm, 0.9*cm,
        f'Gerado em {date.today().strftime("%d/%m/%Y")} — Mobilizze Sistemas')
    canvas.drawRightString(w - 1.8*cm, 0.9*cm,
        f'Página {canvas.getPageNumber()}')

    canvas.restoreState()


def _estilo_tabela_simples(n_colunas=None):
    """Estilo base para todas as tabelas — limpo como Word."""
    return TableStyle([
        # Header
        ('BACKGROUND',    (0, 0), (-1, 0), CINZA_BG),
        ('TEXTCOLOR',     (0, 0), (-1, 0), CINZA_ESC),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 8),
        # Corpo
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('TEXTCOLOR',     (0, 1), (-1, -1), PRETO),
        # Linhas horizontais apenas
        ('LINEBELOW',     (0, 0), (-1, -2), 0.4, CINZA_BORD),
        ('LINEBELOW',     (0, -1), (-1, -1), 0.4, CINZA_BORD),
        ('LINEBEFORE',    (0, 0), (0, -1), 0, BRANCO),
        # Padding
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ])


# ══════════════════════════════════════════════════════════════════════════════
# PDF — DRE Mensal
# ══════════════════════════════════════════════════════════════════════════════
def pdf_dre(contrato, mes, medicoes, gastos_por_categoria,
            fat_bruto, fat_glosa, fat_retencao, fat_liquido,
            total_gastos, margem, margem_pct,
            fat_acumulado, gastos_acumulado,
            margem_acumulada, margem_acumulada_pct):

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2.2*cm, bottomMargin=2*cm
    )

    titulo_s = ParagraphStyle('t', fontSize=16, fontName='Helvetica-Bold',
                               textColor=PRETO, spaceAfter=2)
    sub_s    = ParagraphStyle('s', fontSize=9,  fontName='Helvetica',
                               textColor=CINZA_MED, spaceAfter=16)
    secao_s  = ParagraphStyle('sec', fontSize=8, fontName='Helvetica-Bold',
                               textColor=CINZA_MED, spaceBefore=18, spaceAfter=6,
                               leading=10)

    story = []

    story.append(Paragraph(f'DRE Mensal', titulo_s))
    story.append(Paragraph(
        f'{contrato.numero_sap}  ·  {contrato.empresa.razao_social}  ·  '
        f'{mes.strftime("%m/%Y")}', sub_s))
    story.append(HRFlowable(width='100%', color=CINZA_BORD, thickness=0.5))

    # ── Acumulado ─────────────────────────────────────────────────────────────
    story.append(Paragraph('RESUMO ACUMULADO', secao_s))
    kpi = [
        ['Faturamento', 'Gastos', 'Margem', 'Margem %'],
        [f'R$ {fat_acumulado:,.2f}', f'R$ {gastos_acumulado:,.2f}',
         f'R$ {margem_acumulada:,.2f}', f'{margem_acumulada_pct}%'],
    ]
    t = Table(kpi, colWidths=[4.3*cm]*4)
    s = _estilo_tabela_simples()
    s.add('ALIGN', (0,0), (-1,-1), 'RIGHT')
    s.add('ALIGN', (0,0), (-1,0), 'LEFT')
    t.setStyle(s)
    story.append(t)

    # ── DRE do mês ────────────────────────────────────────────────────────────
    story.append(Paragraph(f'DRE — {mes.strftime("%m/%Y")}', secao_s))

    dre = [['Descrição', 'Valor (R$)']]
    dre.append(['Faturamento bruto', f'R$ {fat_bruto:,.2f}'])
    if fat_glosa > 0:
        dre.append(['(−) Glosas', f'R$ {fat_glosa:,.2f}'])
    if fat_retencao > 0:
        dre.append(['(−) Retenções', f'R$ {fat_retencao:,.2f}'])
    dre.append(['Faturamento líquido', f'R$ {fat_liquido:,.2f}'])

    for cat, valor in gastos_por_categoria.items():
        dre.append([f'(−) {cat}', f'R$ {valor:,.2f}'])
    dre.append(['Total de gastos', f'R$ {total_gastos:,.2f}'])
    dre.append(['Margem bruta', f'R$ {margem:,.2f}  ({margem_pct}%)'])

    t2 = Table(dre, colWidths=[13*cm, 4.5*cm])
    s2 = _estilo_tabela_simples()
    s2.add('ALIGN', (1,0), (1,-1), 'RIGHT')

    # Faturamento líquido em negrito
    liq = 1 + (1 if fat_glosa > 0 else 0) + (1 if fat_retencao > 0 else 0) + 1
    liq_idx = liq - 1
    s2.add('FONTNAME', (0, liq_idx), (-1, liq_idx), 'Helvetica-Bold')

    # Margem em cor
    last = len(dre) - 1
    cor = VERDE if margem >= 0 else VERMELHO
    s2.add('FONTNAME',  (0, last), (-1, last), 'Helvetica-Bold')
    s2.add('TEXTCOLOR', (0, last), (-1, last), cor)

    t2.setStyle(s2)
    story.append(t2)

    # ── Medições ──────────────────────────────────────────────────────────────
    if medicoes:
        story.append(Paragraph('MEDIÇÕES DO MÊS', secao_s))
        med = [['BM', 'Status', 'Valor Bruto', 'Glosa', 'Líquido', 'NF']]
        for m in medicoes:
            med.append([
                f'BM-{m.numero:03d}',
                m.get_status_display(),
                f'R$ {m.valor_bruto:,.2f}',
                f'R$ {m.valor_glosa:,.2f}',
                f'R$ {m.valor_liquido:,.2f}',
                m.numero_nf or '—',
            ])
        t3 = Table(med, colWidths=[2*cm, 2.8*cm, 3*cm, 2.8*cm, 3*cm, 3.9*cm])
        s3 = _estilo_tabela_simples()
        s3.add('ALIGN', (2,1), (4,-1), 'RIGHT')
        t3.setStyle(s3)
        story.append(t3)

    def _h(c, d):
        _header_pdf(c, d,
            titulo=f'DRE — {contrato.numero_sap}',
            subtitulo=mes.strftime('%m/%Y'))

    doc.build(story, onFirstPage=_h, onLaterPages=_h)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# PDF — Quadro de Mobilização
# ══════════════════════════════════════════════════════════════════════════════
def pdf_mobilizacao(contrato, cargos, colaboradores):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2.2*cm, bottomMargin=2*cm
    )

    titulo_s = ParagraphStyle('t', fontSize=16, fontName='Helvetica-Bold',
                               textColor=PRETO, spaceAfter=2)
    sub_s    = ParagraphStyle('s', fontSize=9,  fontName='Helvetica',
                               textColor=CINZA_MED, spaceAfter=16)
    secao_s  = ParagraphStyle('sec', fontSize=8, fontName='Helvetica-Bold',
                               textColor=CINZA_MED, spaceBefore=18, spaceAfter=6)

    story = []
    story.append(Paragraph('Quadro de Mobilização', titulo_s))
    story.append(Paragraph(
        f'{contrato.numero_sap}  ·  {contrato.empresa.razao_social}  ·  '
        f'{date.today().strftime("%d/%m/%Y")}', sub_s))
    story.append(HRFlowable(width='100%', color=CINZA_BORD, thickness=0.5))

    # ── Quadro mínimo ─────────────────────────────────────────────────────────
    story.append(Paragraph('QUADRO MÍNIMO OBRIGATÓRIO', secao_s))
    cargo_data = [['Função', 'CBO', 'Mínimo', 'Ativos', 'Situação']]
    for c in cargos:
        situacao = {
            'ok':      'Regular',
            'alerta':  f'Déficit {c.deficit}',
            'critico': f'Déficit {c.deficit} — crítico',
        }.get(c.situacao, '—')
        cargo_data.append([
            c.funcao + (' ★' if c.is_critico else ''),
            c.cbo or '—',
            str(c.quantidade_minima),
            str(c.total_ativos),
            situacao,
        ])

    t = Table(cargo_data, colWidths=[7*cm, 2*cm, 2*cm, 2*cm, 4.5*cm])
    s = _estilo_tabela_simples()
    s.add('ALIGN', (1,0), (3,-1), 'CENTER')

    for i, c in enumerate(cargos, start=1):
        if c.situacao == 'critico':
            s.add('TEXTCOLOR', (0,i), (-1,i), VERMELHO)
            s.add('FONTNAME',  (0,i), (-1,i), 'Helvetica-Bold')
        elif c.situacao == 'alerta':
            s.add('TEXTCOLOR', (4,i), (4,i), colors.HexColor('#d97706'))

    t.setStyle(s)
    story.append(t)

    # ── Colaboradores ─────────────────────────────────────────────────────────
    story.append(Paragraph('COLABORADORES MOBILIZADOS', secao_s))
    if not colaboradores:
        story.append(Paragraph('Nenhum colaborador mobilizado.',
            ParagraphStyle('v', fontSize=9, textColor=CINZA_MED)))
    else:
        colab = [['Nome', 'CPF', 'Matrícula', 'Função', 'Mobilização', 'Status']]
        for col in colaboradores:
            colab.append([
                col.nome_completo, col.cpf,
                col.matricula_empresa or '—', col.funcao,
                col.data_mobilizacao.strftime('%d/%m/%Y'),
                col.get_status_display(),
            ])
        t2 = Table(colab, colWidths=[4.5*cm, 2.5*cm, 2*cm, 3*cm, 2.5*cm, 3*cm])
        t2.setStyle(_estilo_tabela_simples())
        story.append(t2)

    def _h(c, d):
        _header_pdf(c, d, titulo=f'Mobilização — {contrato.numero_sap}')

    doc.build(story, onFirstPage=_h, onLaterPages=_h)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# PDF — Lista de Contratos
# ══════════════════════════════════════════════════════════════════════════════
def pdf_contratos(contratos):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2.2*cm, bottomMargin=2*cm
    )

    titulo_s = ParagraphStyle('t', fontSize=16, fontName='Helvetica-Bold',
                               textColor=PRETO, spaceAfter=2)
    sub_s    = ParagraphStyle('s', fontSize=9,  fontName='Helvetica',
                               textColor=CINZA_MED, spaceAfter=16)

    story = []
    story.append(Paragraph('Lista de Contratos', titulo_s))
    story.append(Paragraph(
        f'{len(contratos)} contrato(s)  ·  '
        f'{date.today().strftime("%d/%m/%Y")}', sub_s))
    story.append(HRFlowable(width='100%', color=CINZA_BORD, thickness=0.5))
    story.append(Spacer(1, 0.3*cm))

    data = [['Nº SAP', 'Empresa', 'Área', 'Valor (R$)',
             'Início', 'Término', 'Status']]
    for c in contratos:
        data.append([
            c.numero_sap,
            c.empresa.razao_social[:40],
            c.get_area_display(),
            f'R$ {c.valor_atual:,.0f}',
            c.data_inicio.strftime('%d/%m/%Y'),
            c.data_termino_atual.strftime('%d/%m/%Y'),
            c.get_status_display(),
        ])

    t = Table(data, colWidths=[3*cm, 9*cm, 3*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    s = _estilo_tabela_simples()
    s.add('ALIGN', (3,1), (3,-1), 'RIGHT')

    status_cor = {
        'vigente':   VERDE,
        'suspenso':  colors.HexColor('#d97706'),
        'encerrado': CINZA_MED,
    }
    for i, c in enumerate(contratos, start=1):
        cor = status_cor.get(c.status)
        if cor:
            s.add('TEXTCOLOR', (6,i), (6,i), cor)
            s.add('FONTNAME',  (6,i), (6,i), 'Helvetica-Bold')

    t.setStyle(s)
    story.append(t)

    def _h(cv, d):
        _header_pdf(cv, d, titulo='Lista de Contratos')

    doc.build(story, onFirstPage=_h, onLaterPages=_h)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — helpers
# ══════════════════════════════════════════════════════════════════════════════
def _excel_estilos():
    borda_baixo = Border(
        bottom=Side(style='thin', color='dddddd')
    )
    borda_header = Border(
        bottom=Side(style='medium', color='111111')
    )
    return {
        'titulo':  Font(name='Calibri', size=16, bold=True,  color='111111'),
        'sub':     Font(name='Calibri', size=9,  bold=False, color='888888'),
        'header':  Font(name='Calibri', size=9,  bold=True,  color='111111'),
        'normal':  Font(name='Calibri', size=9,  bold=False, color='111111'),
        'bold':    Font(name='Calibri', size=9,  bold=True,  color='111111'),
        'azul':    Font(name='Calibri', size=9,  bold=True,  color='1d4ed8'),
        'verde':   Font(name='Calibri', size=10, bold=True,  color='16a34a'),
        'verm':    Font(name='Calibri', size=10, bold=True,  color='dc2626'),
        'cinza':   Font(name='Calibri', size=9,  bold=False, color='888888'),
        'borda':        borda_baixo,
        'borda_header': borda_header,
        'centro':   Alignment(horizontal='center', vertical='center'),
        'direita':  Alignment(horizontal='right',  vertical='center'),
        'esquerda': Alignment(horizontal='left',   vertical='center'),
        'fmt_brl':  '#,##0.00',
        'fmt_pct':  '0.0"%"',
        'cinza_bg': PatternFill('solid', fgColor='f7f7f7'),
    }


def _excel_cabecalho(ws, titulo, subtitulo, n_colunas, e):
    """Título + subtítulo padrão Word."""
    ws.sheet_view.showGridLines = False

    ultima = get_column_letter(n_colunas)
    ws.merge_cells(f'A1:{ultima}1')
    ws['A1'] = titulo
    ws['A1'].font = e['titulo']
    ws['A1'].alignment = e['esquerda']
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f'A2:{ultima}2')
    ws['A2'] = subtitulo
    ws['A2'].font = e['sub']
    ws['A2'].alignment = e['esquerda']
    ws.row_dimensions[2].height = 18

    # Linha separadora via borda inferior
    for col in range(1, n_colunas + 1):
        ws.cell(2, col).border = Border(
            bottom=Side(style='medium', color='111111')
        )
    ws.row_dimensions[3].height = 10


def _excel_header_row(ws, linha, headers, e):
    for i, h in enumerate(headers, 1):
        c = ws.cell(linha, i, h)
        c.font      = e['header']
        c.alignment = e['esquerda']
        c.border    = e['borda_header']
    ws.row_dimensions[linha].height = 20


def _excel_data_row(ws, linha, vals, e, fmts=None, fonts=None):
    for i, val in enumerate(vals, 1):
        c = ws.cell(linha, i, val)
        c.font      = (fonts[i-1] if fonts and fonts[i-1]
                       else e['normal'])
        c.border    = e['borda']
        c.alignment = (e['direita'] if isinstance(val, float)
                       else e['esquerda'])
        if fmts and fmts[i-1]:
            c.number_format = fmts[i-1]
    ws.row_dimensions[linha].height = 16


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — DRE
# ══════════════════════════════════════════════════════════════════════════════
def excel_dre(contrato, mes, medicoes, gastos_por_categoria,
              fat_bruto, fat_glosa, fat_retencao, fat_liquido,
              total_gastos, margem, margem_pct,
              fat_acumulado, gastos_acumulado,
              margem_acumulada, margem_acumulada_pct):

    wb = Workbook()
    e  = _excel_estilos()

    # ── Aba DRE ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = f'DRE {mes.strftime("%m-%Y")}'
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 22

    _excel_cabecalho(ws, 'DRE Mensal',
        f'{contrato.numero_sap}  ·  {contrato.empresa.razao_social}  ·  '
        f'Competência {mes.strftime("%m/%Y")}', 2, e)

    # Acumulado
    linha = 4
    ws.cell(linha, 1, 'RESUMO ACUMULADO').font = Font(
        name='Calibri', size=8, bold=True, color='888888')
    ws.row_dimensions[linha].height = 18
    linha += 1

    headers_kpi = [
        ('Faturamento acumulado', fat_acumulado,       e['fmt_brl'], None),
        ('Gastos acumulados',     gastos_acumulado,    e['fmt_brl'], None),
        ('Margem acumulada',      margem_acumulada,    e['fmt_brl'], None),
        ('Margem %',              margem_acumulada_pct/100, e['fmt_pct'], None),
    ]
    for desc, val, fmt, _ in headers_kpi:
        cd = ws.cell(linha, 1, desc)
        cv = ws.cell(linha, 2, val)
        cd.font = e['normal']
        cv.font = e['bold']
        cd.border = e['borda']
        cv.border = e['borda']
        cv.number_format = fmt
        cv.alignment = e['direita']
        cd.alignment = e['esquerda']
        ws.row_dimensions[linha].height = 16
        linha += 1

    # Espaço
    ws.row_dimensions[linha].height = 12
    linha += 1

    # DRE
    ws.cell(linha, 1, 'DRE DO MÊS').font = Font(
        name='Calibri', size=8, bold=True, color='888888')
    ws.row_dimensions[linha].height = 18
    linha += 1

    _excel_header_row(ws, linha, ['Descrição', 'Valor (R$)'], e)
    linha += 1

    def _linha_dre(desc, val, font=None, fmt=None):
        nonlocal linha
        cd = ws.cell(linha, 1, desc)
        cv = ws.cell(linha, 2, val)
        cd.font = font or e['normal']
        cv.font = font or e['normal']
        cd.border = e['borda']
        cv.border = e['borda']
        cv.number_format = fmt or e['fmt_brl']
        cv.alignment = e['direita']
        cd.alignment = e['esquerda']
        ws.row_dimensions[linha].height = 16
        linha += 1

    _linha_dre('Faturamento bruto', fat_bruto)
    if fat_glosa > 0:
        _linha_dre('(−) Glosas', fat_glosa,
                   font=Font(name='Calibri', size=9, color='dc2626'))
    if fat_retencao > 0:
        _linha_dre('(−) Retenções', fat_retencao,
                   font=Font(name='Calibri', size=9, color='dc2626'))
    _linha_dre('Faturamento líquido', fat_liquido, font=e['azul'])

    ws.row_dimensions[linha].height = 8
    linha += 1

    for cat, val in gastos_por_categoria.items():
        _linha_dre(f'(−) {cat}', val,
                   font=Font(name='Calibri', size=9, color='dc2626'))
    _linha_dre('Total de gastos', total_gastos,
               font=Font(name='Calibri', size=9, bold=True, color='dc2626'))

    ws.row_dimensions[linha].height = 8
    linha += 1

    _linha_dre(f'Margem bruta  ({margem_pct}%)', margem,
               font=e['verde'] if margem >= 0 else e['verm'])

    # Medições
    if medicoes:
        ws.row_dimensions[linha].height = 16
        linha += 1
        ws.cell(linha, 1, 'MEDIÇÕES DO MÊS').font = Font(
            name='Calibri', size=8, bold=True, color='888888')
        ws.row_dimensions[linha].height = 18
        linha += 1

        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 16

        _excel_header_row(ws, linha,
            ['BM', 'Status', 'Valor Bruto', 'Glosa', 'Líquido', 'NF'], e)
        linha += 1
        for m in medicoes:
            _excel_data_row(ws, linha, [
                f'BM-{m.numero:03d}', m.get_status_display(),
                m.valor_bruto, m.valor_glosa,
                m.valor_liquido, m.numero_nf or '—'
            ], e, fmts=[None, None, e['fmt_brl'], e['fmt_brl'],
                        e['fmt_brl'], None])
            linha += 1

    # ── Aba Gastos ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Gastos por Categoria')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    ws2.sheet_view.showGridLines = False

    _excel_cabecalho(ws2, 'Gastos por Categoria',
        f'{contrato.numero_sap}  ·  {mes.strftime("%m/%Y")}', 2, e)

    _excel_header_row(ws2, 4, ['Categoria', 'Total (R$)'], e)
    for r, (cat, val) in enumerate(gastos_por_categoria.items(), start=5):
        _excel_data_row(ws2, r, [cat, val], e,
                        fmts=[None, e['fmt_brl']])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — Mobilização
# ══════════════════════════════════════════════════════════════════════════════
def excel_mobilizacao(contrato, cargos, colaboradores):
    wb = Workbook()
    e  = _excel_estilos()

    # ── Aba Quadro Mínimo ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Quadro Mínimo'
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([30, 10, 10, 10, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _excel_cabecalho(ws, 'Quadro de Mobilização',
        f'{contrato.numero_sap}  ·  {contrato.empresa.razao_social}  ·  '
        f'{date.today().strftime("%d/%m/%Y")}', 5, e)

    _excel_header_row(ws, 4,
        ['Função / Cargo', 'CBO', 'Mínimo', 'Ativos', 'Situação'], e)

    for r, cargo in enumerate(cargos, start=5):
        situacao = {
            'ok':      'Regular',
            'alerta':  f'Déficit {cargo.deficit}',
            'critico': f'Déficit {cargo.deficit} — crítico',
        }.get(cargo.situacao, '—')
        fonts = [None, None, None, None,
                 Font(name='Calibri', size=9, bold=True, color='dc2626')
                 if cargo.situacao == 'critico' else
                 Font(name='Calibri', size=9, color='d97706')
                 if cargo.situacao == 'alerta' else
                 Font(name='Calibri', size=9, color='16a34a')]
        _excel_data_row(ws, r, [
            cargo.funcao + (' ★' if cargo.is_critico else ''),
            cargo.cbo or '—',
            cargo.quantidade_minima,
            cargo.total_ativos,
            situacao,
        ], e, fonts=fonts)

    # ── Aba Colaboradores ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Colaboradores')
    ws2.sheet_view.showGridLines = False
    for i, w in enumerate([28, 14, 12, 20, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    _excel_cabecalho(ws2, 'Colaboradores Mobilizados',
        f'{contrato.numero_sap}  ·  {contrato.empresa.razao_social}', 6, e)

    _excel_header_row(ws2, 4,
        ['Nome', 'CPF', 'Matrícula', 'Função', 'Mobilização', 'Status'], e)

    status_cor = {
        'mobilizado':    Font(name='Calibri', size=9, color='16a34a'),
        'afastado':      Font(name='Calibri', size=9, color='d97706'),
        'ferias':        Font(name='Calibri', size=9, color='d97706'),
        'desmobilizado': Font(name='Calibri', size=9, color='888888'),
    }
    for r, col in enumerate(colaboradores, start=5):
        fonts = [None, None, None, None, None,
                 status_cor.get(col.status)]
        _excel_data_row(ws2, r, [
            col.nome_completo, col.cpf,
            col.matricula_empresa or '—', col.funcao,
            col.data_mobilizacao.strftime('%d/%m/%Y'),
            col.get_status_display(),
        ], e, fonts=fonts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — Contratos
# ══════════════════════════════════════════════════════════════════════════════
def excel_contratos(contratos):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Contratos'
    ws.sheet_view.showGridLines = False
    e = _excel_estilos()

    for i, w in enumerate([14, 36, 14, 18, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _excel_cabecalho(ws, 'Lista de Contratos',
        f'{len(contratos)} contrato(s)  ·  '
        f'{date.today().strftime("%d/%m/%Y")}', 7, e)

    _excel_header_row(ws, 4,
        ['Nº SAP', 'Empresa', 'Área', 'Valor (R$)',
         'Início', 'Término', 'Status'], e)

    status_cor = {
        'vigente':   Font(name='Calibri', size=9, bold=True, color='16a34a'),
        'suspenso':  Font(name='Calibri', size=9, bold=True, color='d97706'),
        'encerrado': Font(name='Calibri', size=9, color='888888'),
    }
    for r, c in enumerate(contratos, start=5):
        fonts = [None, None, None, None, None, None,
                 status_cor.get(c.status)]
        _excel_data_row(ws, r, [
            c.numero_sap,
            c.empresa.razao_social,
            c.get_area_display(),
            c.valor_atual,
            c.data_inicio.strftime('%d/%m/%Y'),
            c.data_termino_atual.strftime('%d/%m/%Y'),
            c.get_status_display(),
        ], e, fmts=[None, None, None, '#,##0.00',
                    None, None, None],
           fonts=fonts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf